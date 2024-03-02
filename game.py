from openpyxl import Workbook
from strategies import detStrat,randStrat
import time


def conv(choice):
    if choice == "dove":
        return 0
    else:
        return 1


def convback(bit):
    if bit == 0:
        return "dove"
    else:
        return "hawk"


def game(rounds, player1, player2, matrix):  # runs one game of two specific players with a set number of rounds
    moves1 = []
    moves2 = []
    scores = [0, 0]
    if player1[0] == "$":
        choice_oracle1 = getattr(randStrat, player1[1:])
    else:
        choice_oracle1 = getattr(detStrat, player1)

    if player2[0] == "$":
        choice_oracle2 = getattr(randStrat, player2[1:])
    else:
        choice_oracle2 = getattr(detStrat, player2)

    for i in range(rounds):
        choice1 = choice_oracle1(moves1, moves2)
        choice2 = choice_oracle2(moves2, moves1)
        scores[0] += matrix[conv(choice2)][conv(choice1)][1]
        scores[1] += matrix[conv(choice2)][conv(choice1)][0]
        moves1.append(choice1)
        moves2.append(choice2)
    return scores


def full_game(players, rounds, retries, matrix):  # runs whole tourney
    starting = time.time()
    print(f"starting game at {starting}")
    length = len(players)
    scores = [0] * length
    table = [[0 for i in range(length)] for j in range(length)]

    # Initiating excel sheet
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = "Average"
    for i in range(length):
        ws.cell(row=2, column=i + 2).value = players[i]
        ws.cell(row=i + 3, column=1).value = players[i]

    for i in range(length):
        for j in range(i, length):
            s1 = 0
            s2 = 0
            rand1, rand2 = False,False
            player1, player2 = players[i], players[j]
            if player1[0] == "$":
                rand1 = True
                player1 = player1[1:]
            if player2[0] == "$":
                rand2 = True
                player2 = player2[1:]
            if rand1 or rand2:
                for times in range(retries):
                    s = game(rounds, players[i], players[j], matrix)
                    s1 += s[0]
                    s2 += s[1]
            else:
                s = game(rounds, players[i], players[j], matrix)
                s1 += s[0] * retries
                s2 += s[1] * retries

            score1 = s1 / (retries * rounds)
            score2 = s2 / (retries * rounds)
            scores[i] += score1
            if i != j:
                scores[j] += score2
            ws.cell(row=j + 3, column=i + 2).value = score1
            ws.cell(row=i + 3, column=j + 2).value = score2
            # print(f"Game {player1} vs {player2} completed at {time.time()-starting} with score {score1}:{score2}")


    scores = list(a / length for a in scores)
    final = []

    for p in range(length):
        final.append((players[p], round(scores[p], 5)))
    for i in range(length):
        ws.cell(row=1, column=i + 2).value = final[i][1]

    final.sort(key=lambda a: a[1], reverse=True)

    for i in range(length):
        print(f'{final[i][0]} earned {final[i][1]} per round')

    wb.save("results.xlsx")
